import requests
import openpyxl
from bs4 import BeautifulSoup

excel_table = openpyxl.load_workbook('test.xlsx')
sheet = excel_table.worksheets[0]
INN = []
for column in sheet.iter_cols():
	column_name = column[0].value
	if column_name == 'ИНН':
		for cell in column:
			if (cell.value != 'ИНН') and (cell.value != None):
				INN.append(str(cell.value))

sheet.cell(row=1, column=2, value="Название организации")
sheet.cell(row=1, column=3, value="Дата создания")
sheet.cell(row=1, column=4, value="Руководитель")
sheet.cell(row=1, column=5, value="Номер телефона")
sheet.cell(row=1, column=6, value="E-Mail")
sheet.cell(row=1, column=7, value="Юридический адрес")
#Заголовки
headers = {
	'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36 Edg/129.0.0.0'
}

#URL сайта
URL = 'https://vbankcenter.ru'

i = 0
for item in INN:
		
		#URL поисковика на сайте
		SEARCH_URL = URL + '/contragent/search?searchStr=' + INN[i]

		#Главная страница сайта
		main_page = requests.get(SEARCH_URL, headers)
		main_page_soup = BeautifulSoup(main_page.text, 'html.parser')

		#Ссылка на страницу организации
		link_to_org = main_page_soup.find('a', class_ = 'overlap text-blue').get('href')
		URL_ORG = URL + link_to_org

		#Страница организации
		page_org = requests.get(URL_ORG, headers)
		page_org_soup = BeautifulSoup(page_org.text, 'html.parser')

		#Название организации
		title = page_org_soup.find('h1', class_ = 'text-2xl font-semibold mr-2')
		if title == None:
			sheet.cell(row=i+2, column=2, value="------")
		else:
			title = page_org_soup.find('h1', class_ = 'text-2xl font-semibold mr-2').text
			sheet.cell(row=i+2, column=2, value=title)

		#Блок со всеми реквизитами
		req_block = page_org_soup.find('div', class_ = 'requisites-ul-item grid items-start gap-y-4 gap-x-12')
		sections = req_block.findAll('section')

		#Дата создания
		date_block = sections[0]
		date = date_block.find('span')
		if date == None:
			sheet.cell(row=i+2, column=3, value="------")
		else:
			date = date_block.find('span').text

			sheet.cell(row=i+2, column=3, value=date)

		#Руководитель
		director = req_block.find('a', class_ = 'inline-block mb-1 w-max')
		if director == None:
			sheet.cell(row=i+2, column=4, value='------')
		else:
			director = req_block.find('a', class_ = 'inline-block mb-1 w-max').text
			sheet.cell(row=i+2, column=4, value=director)

		#Блок с контактами
		contacts_block = sections[3]

		#Номер телефона
		phone_number = contacts_block.find('span')
		if phone_number == None:
			sheet.cell(row=i+2, column=5, value='------')
		else:
			phone_number = contacts_block.find('span').text
			sheet.cell(row=i+2, column=5, value=phone_number)

		#E-Mail
		e_mail = contacts_block.find('a')
		if e_mail == None:
			sheet.cell(row=i+2, column=6, value='------')
		else:
			e_mail = contacts_block.find('a').text
			sheet.cell(row=i+2, column=6, value=e_mail)

		#Юридический адрес
		adress_block = sections[2]
		adress = adress_block.find('span')
		if adress == None:
			sheet.cell(row=i+2, column=7, value='------')
		else:
			adress = adress_block.find('span').text
			sheet.cell(row=i+2, column=7, value=adress)

		i += 1

excel_table.save('test.xlsx')
